#!/usr/bin/env python3
"""FastAPI server for Mining Globe."""

import os
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse

load_dotenv()

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

MINES = []


def load_xlsx():
    wb = openpyxl.load_workbook("mines.xlsx", read_only=True, data_only=True)
    mines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        country = sheet_name
        current_commodity = None

        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 2 or not row[0]:
                continue
            first = str(row[0]).strip()
            if "No active" in first:
                break
            if first.upper() == first and "(" in first:
                current_commodity = first.split("(")[0].split("\u2014")[0].split("\u2013")[0].strip()
                continue
            if first == "Mine Name":
                continue
            if row[1]:
                lat = row[10] if len(row) > 10 and isinstance(row[10], (int, float)) else None
                lon = row[11] if len(row) > 11 and isinstance(row[11], (int, float)) else None
                if lat is None or lon is None:
                    continue
                mines.append({
                    "country": country,
                    "mine_name": first,
                    "commodity": current_commodity,
                    "operator": str(row[1] or ""),
                    "owning_company": str(row[2] or "") if len(row) > 2 else "",
                    "ticker": str(row[3] or "") if len(row) > 3 else "",
                    "exchange": str(row[4] or "") if len(row) > 4 else "",
                    "mine_type": str(row[5] or "") if len(row) > 5 else "",
                    "annual_production": str(row[6] or "") if len(row) > 6 else "",
                    "status": str(row[7] or "") if len(row) > 7 else "",
                    "location": str(row[8] or "") if len(row) > 8 else "",
                    "notes": str(row[9] or "") if len(row) > 9 else "",
                    "lat": lat,
                    "lon": lon,
                })
    wb.close()
    return mines


@app.on_event("startup")
def startup():
    global MINES
    MINES = load_xlsx()
    print(f"Loaded {len(MINES)} mines")


@app.get("/", response_class=HTMLResponse)
def serve_frontend():
    cesium_token = os.getenv("CESIUM_ION_TOKEN", "")
    html = Path("index.html").read_text()
    return HTMLResponse(content=html.replace("{{CESIUM_ION_TOKEN}}", cesium_token))


@app.get("/api/mines")
def get_mines(
    commodity: str = Query(default=None),
    country: str = Query(default=None),
    status: str = Query(default=None),
    search: str = Query(default=None),
):
    results = MINES
    if commodity:
        c = commodity.upper()
        results = [m for m in results if m["commodity"] and c in m["commodity"].upper()]
    if country:
        results = [m for m in results if m["country"] == country]
    if status:
        results = [m for m in results if m["status"] == status]
    if search:
        s = search.lower()
        results = [m for m in results if s in m["mine_name"].lower() or s in m["operator"].lower() or s in m["owning_company"].lower() or s in m["ticker"].lower()]
    return results


@app.get("/api/stats")
def get_stats():
    commodities = sorted(set(m["commodity"] for m in MINES if m["commodity"]))
    countries = sorted(set(m["country"] for m in MINES))
    return {
        "total_mines": len(MINES),
        "countries": len(countries),
        "commodity_list": commodities,
        "country_list": countries,
    }


if __name__ == "__main__":
    import uvicorn
    print("Starting Mining Globe at http://localhost:8001")
    uvicorn.run(app, host="0.0.0.0", port=8001)
